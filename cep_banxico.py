# cep_banxico.py
from __future__ import annotations
# ————————————————————————————————————————————————————————————————
# CEP Banxico: leer Excel de movimientos, detectar candidatos y
# automatizar el formulario (en iframe o main) para descargar PDFs.
# Requisitos:
#   pip install playwright pandas openpyxl
#   playwright install --with-deps chromium
# ————————————————————————————————————————————————————————————————

import sys, asyncio
if sys.platform.startswith("win"):
    try:
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    except Exception:
        pass

import unicodedata, re
import os, re, json, zipfile, shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# -------------------------------------------------------------------
# Extractores opcionales si se usa modo --pdf
# -------------------------------------------------------------------
try:
    from santander_extractor import extract_santander_to_xlsx
except Exception:
    extract_santander_to_xlsx = None

try:
    from bbva_extractor import extract_bbva_to_xlsx
except Exception:
    extract_bbva_to_xlsx = None

try:
    from banorte_extractor import extract_banorte_to_xlsx
except Exception:
    extract_banorte_to_xlsx = None

try:
    from inbursa_extractor import extract_inbursa_to_xlsx
except Exception:
    extract_inbursa_to_xlsx = None


# -------------------------------------------------------------------
# Diccionarios
# -------------------------------------------------------------------
BANKS_BY_CODE: Dict[str, str] = {
    "002":"BANAMEX","012":"BBVA","014":"SANTANDER","021":"HSBC","036":"INBURSA",
    "044":"SCOTIABANK","058":"BANREGIO","059":"INVEX","062":"AFIRME","072":"BANORTE",
    "127":"AZTECA","137":"ACTINVER","143":"BANJERCITO","145":"BANSI","146":"CIBANCO",
    "147":"BANCOPPEL","150":"MULTIVA","154":"MIFEL","166":"BANCO DEL BIENESTAR",
}
OWN_BANK_TO_CODE = {
    "SANTANDER":"014","BBVA":"012","BANORTE":"072",
    "BANORTE (BANCO MERCANTIL DEL NORTE)":"072","BANORTE IXE":"072",
}

# -------------------------------------------------------------------
# Patrones
# -------------------------------------------------------------------
RE_SPEI = re.compile(r"(?i)\bSPEI\b")
RE_CLABE = re.compile(r"\b(\d{18})\b")
# Banxico + Banorte: "CLAVE DE RASTREO:" | "CVE RASTREO:" | "CVE RAST:"
RE_CLAVE_RASTREO = re.compile(
    r"(?is)\b(?:CLAVE\s*DE\s*RASTREO|CLV?E?\s*RAST(?:REO)?|CVE\s*RASTREO|CVE\s*RAST)\b\s*[:\-]?\s*([A-Z0-9\-]{6,})"
)
RE_REF = re.compile(r"(?i)\bREF(?:ERENCIA)?[:\s-]*([A-Z0-9\-]{5,})")
RE_AMOUNT = re.compile(r"\$?\s*\d{1,3}(?:,\d{3})*\.\d{2}")
# BBVA: a veces imprime CLABE con 20 dígitos (00 + 18 reales)
RE_CLABE_BBVA_20 = re.compile(r"(?<!\d)00(\d{18})(?!\d)")
RE_LONG_DIGITS = re.compile(r"\d{18,}")  # apoyo para casos “pegados” de muchos dígitos
# BBVA: CLABE de 20 (00 + 18) seguida de clave; y fallback 18 + clave
RE_BBVA_CLABE20_PLUS_KEY = re.compile(r"(?<!\d)00(\d{18})(?!\d)\s*([A-Z0-9]{10,32})")
RE_CLABE18_PLUS_KEY      = re.compile(r"(?<!\d)(\d{18})(?!\d)\s*([A-Z0-9]{10,32})")

# Palabras clave típicas en BBVA
RE_BBVA_ENVIADO  = re.compile(r"(?i)\bENVIAD[OA]\b")
RE_BBVA_RECIBIDO = re.compile(r"(?i)\bRECIBID[OA]\b")


# Variantes de columnas por hoja (aceptamos listas)
SHEET_SIGNATURES: Dict[str, Dict[str, List[str]]] = {
    # Santander
    "movimientos_cuenta": {
        "date_col": ["Fecha", "Fecha Operación"],
        "desc_col": ["Descripción", "Descripcion"],
        "dep_cols": ["Depósitos", "Depositos", "Abonos"],
        "ret_cols": ["Retiro", "Retiros", "Cargos"],
    },
    "movimientos_inversion": {
        "date_col": ["Fecha", "Fecha Operación"],
        "desc_col": ["Descripción", "Descripcion"],
        "dep_cols": ["Depósitos", "Depositos", "Abonos"],
        "ret_cols": ["Retiro", "Retiros", "Cargos"],
    },
    # BBVA
    "movimientos": {
        "date_col": ["Fecha Operación", "Fecha"],
        "desc_col": ["Descripción", "Descripcion"],
        "dep_cols": ["Abonos","Depósitos","Depositos"],
        "ret_cols": ["Cargos","Retiros","Retiro"],
    },
    # Banorte
    "cuenta_01": {
        "date_col": ["Fecha","Fecha Operación"],
        "desc_col": ["Descripción","Descripcion"],
        "dep_cols": ["Depósitos/Abonos","Depósitos","Depositos","Abonos"],
        "ret_cols": ["Retiros/Cargos","Retiros","Cargos","Retiro"],
    },
    "cuenta_02": {
        "date_col": ["Fecha","Fecha Operación"],
        "desc_col": ["Descripción","Descripcion"],
        "dep_cols": ["Depósitos/Abonos","Depósitos","Depositos","Abonos"],
        "ret_cols": ["Retiros/Cargos","Retiros","Cargos","Retiro"],
    },
}

# -------------------------------------------------------------------
# Utils
# -------------------------------------------------------------------

def _norm(s: str | None) -> str:
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^A-Z0-9 ]", " ", s.upper()).strip()

def _norm(s: str | None) -> str:
    import unicodedata, re
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s.upper().strip()

def _first_existing(df: pd.DataFrame, options: List[str]) -> Optional[str]:
    for c in options:
        if c in df.columns:
            return c
    return None

def _to_float(x) -> float:
    try:
        if pd.isna(x): return 0.0
        return float(str(x).replace("$","").replace(",","").strip())
    except Exception:
        return 0.0

def _norm_date_ddmmyyyy(s: str) -> Optional[str]:
    s = str(s or "").strip()
    m = re.match(r"^(\d{2})[-/](\d{2})[-/](\d{2,4})$", s)
    if not m: return None
    d,M,y = m.groups()
    y4 = y if len(y)==4 else f"20{y}"
    return f"{d}-{M}-{y4}"

def _bank_name_from_code(code3: Optional[str]) -> Optional[str]:
    if not code3: return None
    return BANKS_BY_CODE.get(code3)

def _get_col_like(df: pd.DataFrame, rx: re.Pattern) -> Optional[str]:
    for c in df.columns:
        if rx.search(str(c)):
            return c
    return None

# --- arriba ya están BANKS_BY_CODE y _norm ---

def _bank_code_from_text(text: str | None) -> str | None:
    """
    Intenta obtener el código Banxico (3 dígitos) a partir de texto.
    - Primero busca un código de 3 dígitos explícito.
    - Luego intenta por nombre del banco (usa BANKS_BY_CODE).
    - Incluye sinónimos comunes (BBVA/BANCOMER, SCOTIA/SCOTIABANK, etc.).
    """
    if not text:
        return None
    t = _norm(str(text))

    # 1) código de 3 dígitos
    m = re.search(r"\b(\d{3})\b", t)
    if m and m.group(1) in BANKS_BY_CODE:
        return m.group(1)

    # 2) por nombre exacto de BANKS_BY_CODE
    for code, name in BANKS_BY_CODE.items():
        if _norm(name) in t:
            return code

    # 3) sinónimos habituales
    aliases = {
        "BANCOMER": "012",
        "BBVA": "012",
        "SANTANDER": "014",
        "BANORTE": "072",
        "SCOTIA": "044",
        "SCOTIABANK": "044",
        "HSBC": "021",
        "BANREGIO": "058",
        "INBURSA": "036",
        "AZTECA": "127",
        "BANAMEX": "002",  # aparece mucho en descripciones
        "CITIBANAMEX": "002",
        "STP": "846",
    }
    for key, code in aliases.items():
        if key in t:
            return code

    return None

def _first_other_bank_name_in_text(text: str, own_bank_name: Optional[str]) -> Optional[str]:
    """
    Devuelve el primer banco (por nombre) encontrado en el texto que NO sea tu propio banco.
    Útil para frases tipo 'SPEI ENVIADO BANORTE' en BBVA.
    """
    if not text:
        return None
    T = str(text).upper()
    for name in BANKS_BY_CODE.values():
        if own_bank_name and name == own_bank_name:
            continue
        # normalizamos algunos alias comunes
        aliases = {
            "BBVA": ["BBVA", "BBVA MEXICO", "BBVA MÉXICO", "BBVA BANCOMER"],
            "BANORTE": ["BANORTE", "BANORTE IXE", "BANORTE (BANCO MERCANTIL DEL NORTE)"],
            "SANTANDER": ["SANTANDER"],
            "INBURSA": ["INBURSA", "BANCO INBURSA"],
        }.get(name, [name])
        if any(alias.upper() in T for alias in aliases):
            return name
    return None

def _code_from_bank_in_info(info_df: pd.DataFrame) -> tuple[str | None, str | None]:
    """
    Devuelve (own_bank_code, own_bank_name) a partir de la hoja 'info'.
    Considera tanto 'Banco' escrito con nombre (INBURSA, BBVA, etc.) como
    la presencia de un código en el texto.
    """
    own_code, own_name = None, None
    try:
        raw = str(info_df.iloc[0].get("Banco", "") or "")
        own_code = _bank_code_from_text(raw)
        if own_code:
            own_name = BANKS_BY_CODE.get(own_code)
        else:
            # último intento por coincidencia exacta del nombre en el texto
            r = _norm(raw)
            for code, name in BANKS_BY_CODE.items():
                if _norm(name) in r:
                    own_code, own_name = code, name
                    break
    except Exception:
        pass
    return own_code, own_name

# === NUEVO: detectar la CLABE desde la hoja "cuenta"/"cuentas" =================
def _read_clabe_from_cuenta_sheet(xlsx_path: str) -> Optional[str]:
    try:
        xls = pd.ExcelFile(xlsx_path)
    except Exception:
        return None

    # Buscar hoja cuyo nombre empiece con "cuent" (cuenta, cuentas…)
    cuenta_sheet = None
    for s in xls.sheet_names:
        if re.search(r"^cuent", s.strip().lower()):
            cuenta_sheet = s
            break
    if not cuenta_sheet:
        return None

    try:
        df = pd.read_excel(xlsx_path, sheet_name=cuenta_sheet, dtype=str)
    except Exception:
        return None

    # 1) Si existe columna CLABE explícita
    clabe_col = _get_col_like(df, re.compile(r"\bclabe\b", re.I))
    found: List[str] = []
    if clabe_col:
        for v in df[clabe_col].astype(str).tolist():
            m = re.search(r"(\d{18})", str(v))
            if m:
                found.append(m.group(1))

    # 2) Si no, buscar 18 dígitos en todo el DF
    if not found:
        for col in df.columns:
            for v in df[col].astype(str).tolist():
                m = re.search(r"(\d{18})", str(v))
                if m:
                    found.append(m.group(1))

    found = list(dict.fromkeys(found))  # únicos, preservando orden
    if len(found) == 1:
        return found[0]
    # Si hay varias y no podemos decidir, no forzamos (que caiga al de info)
    return None

import itertools

def _sheet_index_from_name(sheet_name: str) -> Optional[int]:
    """
    Devuelve 0 para 'cuenta_01', 1 para 'cuenta_02', etc. Si no matchea, None.
    Acepta variantes: 'cuenta 01', 'CUENTA_02', etc.
    """
    m = re.match(r"(?i)^\s*cuenta[\s_]*0?(\d+)\s*$", sheet_name or "")
    if not m:
        return None
    idx = int(m.group(1)) - 1
    return idx if idx >= 0 else None

def _read_all_clabes_from_cuentas(xlsx_path: str) -> List[str]:
    """
    Lee la hoja cuyo nombre empiece con 'cuent' (cuenta/cuentas)
    y devuelve todas las CLABEs (18 dígitos) que encuentre por filas,
    en orden de aparición y sin duplicados.
    """
    try:
        xls = pd.ExcelFile(xlsx_path)
    except Exception:
        return []

    cuentas_sheet = None
    for s in xls.sheet_names:
        if re.search(r"^cuent", s.strip().lower()):
            cuentas_sheet = s
            break
    if not cuentas_sheet:
        return []

    try:
        df = pd.read_excel(xlsx_path, sheet_name=cuentas_sheet, dtype=str)
    except Exception:
        return []

    # 1) Si hay columna 'Clabe' explícita
    clabe_col = _get_col_like(df, re.compile(r"\bclabe\b", re.I))
    values: List[str] = []
    if clabe_col:
        for v in df[clabe_col].astype(str).tolist():
            m = re.search(r"(\d{18})", v)
            if m: values.append(m.group(1))

    # 2) Si no, buscar 18 dígitos en todas las columnas
    if not values:
        for v in itertools.chain.from_iterable(df.astype(str).values.tolist()):
            m = re.search(r"(\d{18})", str(v))
            if m: values.append(m.group(1))

    # Únicos preservando orden
    seen, result = set(), []
    for c in values:
        if c not in seen:
            seen.add(c)
            result.append(c)
    return result

def _build_sheet_clabe_map(xlsx_path: str) -> Dict[str, str]:
    """
    Mapea nombre de hoja -> CLABE propia.
    Ej.: {'cuenta_01': CLABE fila1, 'cuenta_02': CLABE fila2}
    """
    clabes = _read_all_clabes_from_cuentas(xlsx_path)
    if not clabes:
        return {}

    mapping: Dict[str, str] = {}
    try:
        xls = pd.ExcelFile(xlsx_path)
    except Exception:
        return {}

    for s in xls.sheet_names:
        idx = _sheet_index_from_name(s)
        if idx is not None and idx < len(clabes):
            mapping[s] = clabes[idx]
        # También soporta alias 'cuenta 01', 'cuenta_01', etc.
    return mapping



# -------------------------------------------------------------------
# Jobs
# -------------------------------------------------------------------
@dataclass
class CepJob:
    sheet: str
    row_index: int
    fecha: str
    monto: float
    clave_rastreo: Optional[str] = None
    numero_referencia: Optional[str] = None
    banco_emisor: Optional[str] = None
    banco_receptor: Optional[str] = None
    cuenta_beneficiaria: Optional[str] = None  # CLABE (18d)

# -------------------------------------------------------------------
# Lectura info de la hoja "info" (robusta) + CLABE desde "cuenta"
# -------------------------------------------------------------------
def _read_info(xlsx_path: str) -> Tuple[Optional[str], Optional[str], Dict[str, str]]:
    """Regresa (own_bank_code3, own_clabe, row_dict). Prefiere CLABE de hoja cuenta."""
    # 1) Primero intentamos hoja "cuenta"
    clabe_from_cuenta = _read_clabe_from_cuenta_sheet(xlsx_path)

    # 2) Luego leemos "info" para el banco y, si falla lo anterior, la CLABE
    try:
        df = pd.read_excel(xlsx_path, sheet_name="info", dtype=str)
    except Exception:
        return None, clabe_from_cuenta, {}

    if df.empty:
        return None, clabe_from_cuenta, {}

    row = df.iloc[0].to_dict()
    banco_col = _get_col_like(df, re.compile(r"banco", re.I)) or "Banco"
    clabe_col = _get_col_like(df, re.compile(r"clabe", re.I)) or "CLABE"
    own_code = _bank_code_from_text(row.get(banco_col))


    # Si no trajimos CLABE de "cuenta", tomamos la de "info"
    if clabe_from_cuenta:
        own_clabe = clabe_from_cuenta
    else:
        raw_clabe = str(row.get(clabe_col) or "").strip()
        m = re.search(r"(\d{18})", raw_clabe)
        own_clabe = m.group(1) if m else None

    row = {k: (None if pd.isna(v) else v) for k, v in row.items()}
    return own_code, own_clabe, row

def _find_signature_for_sheet(df: pd.DataFrame, sheet_name: str) -> Optional[Dict[str, List[str]]]:
    if sheet_name in SHEET_SIGNATURES:
        return SHEET_SIGNATURES[sheet_name]
    cols = set(df.columns)
    if {"Fecha Operación","Descripción"}.issubset(cols) and ({"Abonos","Cargos"} & cols):
        return SHEET_SIGNATURES["movimientos"]
    return None

def _extract_keys_from_desc(
    desc: str,
    own_bank_code: Optional[str] = None
) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Regresa (clave_rastreo, numero_referencia, clabe_en_texto).
    Para BBVA:
      - Normaliza CLABE impresa como '00' + 18.
      - Si la clave viene inmediatamente después de la CLABE, también la extrae.
    """
    if not desc:
        return None, None, None

    t = str(desc)

    # 1) Clave por etiqueta (genérico: Banxico/Banorte/etc.)
    m1 = RE_CLAVE_RASTREO.search(t)
    clave = m1.group(1).replace(" ", "").upper().strip("-") if m1 else None

    # 2) Referencia (opcional)
    m2 = RE_REF.search(t)
    ref = m2.group(1).replace(" ", "").upper().strip("-") if m2 else None

    clabe: Optional[str] = None

    # 3) Reglas especiales BBVA
    if own_bank_code == "012":
        # 3.a) Caso ideal: 00 + CLABE(18) seguido de CLAVE
        m20k = RE_BBVA_CLABE20_PLUS_KEY.search(t)
        if m20k:
            clabe = m20k.group(1)
            if not clave:
                clave = m20k.group(2).upper()

        # 3.b) Fallback: bloque de 20 dígitos "00 + 18"
        if not clabe:
            m20 = RE_CLABE_BBVA_20.search(t)
            if m20:
                clabe = m20.group(1)

        # 3.c) Si hay CLABE(18) seguida de CLAVE, también sirve
        if not clave:
            m18k = RE_CLABE18_PLUS_KEY.search(t)
            if m18k:
                if not clabe:
                    clabe = m18k.group(1)
                clave = m18k.group(2).upper()

        # 3.d) Último respaldo: tiras largas de dígitos empezando en '00'
        if not clabe:
            for mm in RE_LONG_DIGITS.finditer(t):
                s = mm.group(0)
                if len(s) == 20 and s.startswith("00"):
                    clabe = s[2:]
                    break

    # 4) Genérico: si aún no tenemos CLABE, busca una normal de 18
    if not clabe:
        m3 = RE_CLABE.search(t)
        clabe = m3.group(1) if m3 else None

    return clave, ref, clabe

def _detect_jobs_in_sheet(
    df: pd.DataFrame,
    sheet_name: str,
    own_bank_code: Optional[str],
    own_clabe: Optional[str],
    sheet_clabe_map: Optional[Dict[str, str]] = None,
) -> List[CepJob]:
    sig = _find_signature_for_sheet(df, sheet_name)
    if not sig: return []

    date_col = _first_existing(df, sig["date_col"])
    desc_col = _first_existing(df, sig["desc_col"])
    dep_col  = _first_existing(df, sig["dep_cols"])
    ret_col  = _first_existing(df, sig["ret_cols"])
    if not (date_col and desc_col and dep_col and ret_col):
        return []

    own_bank_name = _bank_name_from_code(own_bank_code) if own_bank_code else None

    jobs: List[CepJob] = []
    for i, r in df.iterrows():
        desc = str(r.get(desc_col) or "").strip()
        if not desc:
            continue

        clave, ref, clabe_en_texto = _extract_keys_from_desc(desc, own_bank_code)
        if not (clave or ref):
            if not RE_SPEI.search(desc):
                continue

        fecha = _norm_date_ddmmyyyy(r.get(date_col))
        if not fecha:
            continue

        dep = _to_float(r.get(dep_col))
        ret = _to_float(r.get(ret_col))
        # Determinamos tipo por montos; si empatan o son 0, usamos palabras clave BBVA
        kw_env = bool(RE_BBVA_ENVIADO.search(desc))
        kw_rec = bool(RE_BBVA_RECIBIDO.search(desc))

        is_retiro = ret > dep
        is_deposito = dep > ret

        if not (is_retiro or is_deposito):
            # Desempate por texto (BBVA suele poner 'SPEI ENVIADO/RECIBIDO')
            if kw_env and not kw_rec:
                is_retiro = True
            elif kw_rec and not kw_env:
                is_deposito = True
            else:
                # si no hay pista: tomamos depósito si hay 'dep' o retiro si hay 'ret'
                is_deposito = dep > 0
                is_retiro = ret > 0

        monto = dep if is_deposito else ret
        if monto <= 0:
            mamt = RE_AMOUNT.findall(desc)
            if mamt:
                try:
                    monto = abs(float(mamt[-1].replace("$","").replace(",","")))
                except Exception:
                    pass
        if monto <= 0:
            continue

        # Banco detectado por nombre/código en la descripción
        code_in_desc = _bank_code_from_text(desc)
        bank_from_text = _bank_name_from_code(code_in_desc) if code_in_desc else None

        if is_retiro:
            # RETIRO: tu banco es emisor; receptor = otro banco del texto o por CLABE
            banco_emisor = own_bank_name

            receptor_code = (clabe_en_texto or "")[:3] if clabe_en_texto else code_in_desc
            banco_receptor = _bank_name_from_code(receptor_code) if receptor_code else bank_from_text

            # En BBVA conviene preferir el banco "distinto a BBVA" que aparece después de 'ENVIADO'
            if own_bank_code == "012":
                other = _first_other_bank_name_in_text(desc, own_bank_name)
                if other:
                    banco_receptor = other

            cuenta_beneficiaria = clabe_en_texto  # viene en la descripción (ya normalizada si BBVA)

        else:
            # DEPÓSITO: receptor = tu banco; emisor = banco del texto/CLABE
            emisor_code = (clabe_en_texto or "")[:3] if clabe_en_texto else code_in_desc
            banco_emisor = _bank_name_from_code(emisor_code) if emisor_code else bank_from_text
            banco_receptor = own_bank_name

            # CLABE beneficiaria es la propia (mapeo por hoja si existe; fallback 'info')
            cuenta_beneficiaria = ((sheet_clabe_map or {}).get(sheet_name) or own_clabe)

        jobs.append(CepJob(
            sheet=sheet_name,
            row_index=i,
            fecha=fecha,
            monto=round(abs(monto), 2),
            clave_rastreo=clave,
            numero_referencia=ref,
            banco_emisor=banco_emisor,
            banco_receptor=banco_receptor,
            cuenta_beneficiaria=cuenta_beneficiaria,
        ))
    return jobs

def collect_jobs_from_xlsx(xlsx_path: str) -> Tuple[List[CepJob], Dict[str, List[int]]]:
    own_code, own_clabe, _info = _read_info(xlsx_path)
    sheet_clabe_map = _build_sheet_clabe_map(xlsx_path)
    xls = pd.ExcelFile(xlsx_path)
    all_jobs: List[CepJob] = []
    rows_idx_by_sheet: Dict[str, List[int]] = {}

    for sheet in xls.sheet_names:
        df = pd.read_excel(xlsx_path, sheet_name=sheet)
        jobs = _detect_jobs_in_sheet(df, sheet, own_code, own_clabe, sheet_clabe_map)
        if jobs:
            all_jobs.extend(jobs)
            rows_idx_by_sheet.setdefault(sheet, []).extend([j.row_index for j in jobs])
    return all_jobs, rows_idx_by_sheet

# -------------------------------------------------------------------
# Playwright helpers
# -------------------------------------------------------------------
SPANISH_MONTHS = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12
}
SPANISH_MONTHS_ABBR = {  # primeras 3 letras en minúsculas
    1:"ene",2:"feb",3:"mar",4:"abr",5:"may",6:"jun",7:"jul",8:"ago",9:"sep",10:"oct",11:"nov",12:"dic"
}

def _force_set_input_value(locator, value: str):
    try:
        locator.evaluate(
            "(el, v) => {"
            "  el.value = v;"
            "  el.dispatchEvent(new Event('input', {bubbles:true}));"
            "  el.dispatchEvent(new Event('change', {bubbles:true}));"
            "}", value
        )
    except Exception:
        pass
    try:
        locator.evaluate(
            """
            (el, v) => {
              try {
                if (window.jQuery) {
                  const $ = window.jQuery;
                  if ($(el).data('datepicker')) { $(el).datepicker('setDate', v); $(el).trigger('change'); }
                  const dtp = $(el).data('DateTimePicker');
                  if (dtp && dtp.date) { dtp.date(v); $(el).trigger('change'); }
                  if ($.datepicker && $(el).hasClass('hasDatepicker')) {
                    try { $(el).datepicker('setDate', v); } catch(e) {}
                  }
                }
              } catch(e) {}
            }
            """, value
        )
    except Exception:
        pass

def _handle_captcha(frm, page, workdir: Path, headless: bool) -> bool:
    import re, os
    code = frm.get_by_label(re.compile(r"C[oó]digo\s+de\s+seguridad", re.I))
    if not code or code.count() == 0:
        return True
    try:
        img = frm.locator("img[alt*='caracteres' i], img[src*='captcha' i], img[alt*='seguridad' i]").first
        if img and img.count():
            workdir.mkdir(parents=True, exist_ok=True)
            img.screenshot(path=str(workdir / "_captcha.png"))
            print(f"[CEP] Captcha capturado en: {workdir / '_captcha.png'}")
    except Exception:
        pass
    env_val = os.getenv("CEP_CAPTCHA", "").strip()
    if env_val:
        try:
            code.fill(env_val)
            frm.wait_for_timeout(350)
            return True
        except Exception:
            pass
    if headless:
        print("[CEP] Captcha detectado en modo headless y CEP_CAPTCHA no configurado -> saltando job.")
        return False
    try:
        cur = ""
        try: cur = code.input_value(timeout=600).strip()
        except Exception: pass
        if not cur:
            print("[CEP] Escribe el 'Código de seguridad' en la pestaña del navegador y presiona ENTER aquí…")
            try: input()
            except Exception: pass
        frm.wait_for_timeout(500)
    except Exception:
        pass
    return True

# --- helper para clickear "Descargar CEP" y bajar el PDF ----------------
def _click_descargar_y_bajar_pdf(page, frm, out_path: Path) -> Optional[Path]:
    """
    Minimal (con TAB largos):
      1) Click en 'Descargar CEP'
      2) Espera 4s
      3) Envía 4x TAB (delay largo) + espera 1s + ENTER
      4) Captura la descarga vía expect_download
    """
    import re as _re
    try:
        out_path.parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass

    try:
        btn = frm.locator("#btn_Descargar").first
        if not btn or btn.count() == 0:
            btn = frm.get_by_role("button", name=_re.compile(r"Descargar\s+CEP", _re.I)).first
        if not btn or btn.count() == 0:
            btn = frm.locator("button:has-text('Descargar CEP'), a:has-text('Descargar CEP')").first
    except Exception:
        btn = None

    if not btn or btn.count() == 0:
        try: page.screenshot(path=str(out_path.parent / "debug_btn_no_encontrado.png"), full_page=True)
        except Exception: pass
        return None

    try:
        page.get_by_role("button", name=_re.compile(r"Aceptar|De acuerdo|OK", _re.I)).click(timeout=1200)
    except Exception:
        pass

    try: btn.scroll_into_view_if_needed(timeout=4000)
    except Exception: pass
    try: btn.click(force=True, timeout=4000)
    except Exception:
        try:
            btn.evaluate("el => { el.removeAttribute && el.removeAttribute('disabled'); el.click && el.click(); }")
        except Exception:
            pass

    page.wait_for_timeout(4000)

    TAB_HOLD_MS = 700
    BETWEEN_TABS_MS = 300
    AFTER_TABS_MS = 1000

    try:
        with page.expect_download(timeout=30000) as dlev:
            for _ in range(4):
                page.keyboard.press("Tab", delay=TAB_HOLD_MS)
                page.wait_for_timeout(BETWEEN_TABS_MS)
            page.wait_for_timeout(AFTER_TABS_MS)
            page.keyboard.press("Enter", delay=120)
        d = dlev.value
        d.save_as(str(out_path))
        print("[CEP] PDF descargado (TABx4 largos + ENTER) ->", out_path)
        return out_path
    except Exception as e:
        print("[CEP] No se capturó descarga con TAB largos + ENTER:", repr(e))
        try: page.screenshot(path=str(out_path.parent / "debug_tab_enter_sin_descarga.png"), full_page=True)
        except Exception: pass
        return None

def _parse_header_month_year(txt: str) -> Tuple[Optional[int], Optional[int]]:
    if not txt:
        return None, None
    t = txt.strip().lower()
    parts = re.findall(r"[a-záéíóúñ]+|\d{4}", t)
    if not parts:
        return None, None
    month = None
    year = None
    for p in parts:
        if p.isdigit() and len(p) == 4:
            year = int(p)
        else:
            month = SPANISH_MONTHS.get(p, month)
    return month, year

def _datepicker_container(frame):
    sel = (
        "div.datepicker-dropdown:visible, "
        "div.ui-datepicker:visible, "
        "div.datepicker:visible, "
        "div[role='dialog'] div.datepicker:visible, "
        "div[aria-label*='calendar' i]:visible"
    )
    return frame.locator(sel).first

def _norm_txt(s: Optional[str]) -> str:
    if not s:
        return ""
    import unicodedata, re
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s).strip().upper()
    s = re.sub(r"[^A-Z0-9 ]", "", s)
    return s

_BANK_SYNONYMS = {
    "BBVA": ["BBVA", "BBVA MEXICO", "BBVA MÉXICO", "BBVA BANCOMER"],
    "BANORTE": ["BANORTE", "BANORTE IXE", "BANORTE (BANCO MERCANTIL DEL NORTE)"],
    "SANTANDER": ["SANTANDER"],
    "HSBC": ["HSBC"],
    "INBURSA": ["INBURSA", "BANCO INBURSA"],
}

def _norm_txt(s: Optional[str]) -> str:
    if not s:
        return ""
    import unicodedata, re
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s).strip().upper()
    s = re.sub(r"[^A-Z0-9 ]", "", s)
    return s

_BANK_SYNONYMS_NORM = {
    _norm_txt(k): [_norm_txt(alias) for alias in v]
    for k, v in _BANK_SYNONYMS.items()
}

# Script robusto para elegir una opción por texto (con aliases)
_JS_PICK_OPTION = r"""
function __norm(s){
  return (s||"").normalize('NFD').replace(/[\u0300-\u036f]/g,'')
           .toUpperCase().replace(/[^A-Z0-9 ]/g,'').replace(/\s+/g,' ').trim();
}
return (function(selectEl, wanted, aliases){
  const desired = __norm(wanted);
  const wanteds = new Set([desired].concat(aliases||[]));
  // 1) igualdad exacta o por inclusión amplia
  for (let i = 0; i < selectEl.options.length; i++){
    const t = __norm(selectEl.options[i].textContent);
    if (wanteds.has(t) || t === desired || t.startsWith(desired) || t.includes(" "+desired+" ") || t.endsWith(" "+desired)){
      selectEl.selectedIndex = i;
      selectEl.dispatchEvent(new Event('change', {bubbles:true}));
      return true;
    }
  }
  return false;
})(arguments[0], arguments[1], arguments[2]);
"""
def _get_selected_text(sel) -> str:
    try:
        return (sel.evaluate(
            """(el)=>{const o=el.options[el.selectedIndex];return (o && o.textContent)||'';}"""
        ) or "").strip()
    except Exception:
        return ""

def _select_option_by_label_loose(frame, label_rx: re.Pattern, wanted_text: Optional[str], role_hint: Optional[str] = None) -> None:
    """
    Selección robusta del banco en el <select> de Banxico.
    Soporta ids: cmbInstitucionReceptora/Emisora, ...Beneficiaria/Ordenante, variaciones y etiquetas.
    """
    if not wanted_text:
        return

    desired = _norm_txt(wanted_text)
    aliases_norm = set(_BANK_SYNONYMS_NORM.get(desired, []))

    # 1) Candidatos: por label y por heurística de nombre/id (muy amplio)
    candidates = []

    try:
        by_label = frame.get_by_label(label_rx)
        if by_label and by_label.count():
            candidates.append(by_label)
    except Exception:
        pass

    # claves por rol
    role = (role_hint or "").lower()
    keys = []
    if role.startswith("emis"):
        keys = ["emis", "emiso", "orden", "ordenan"]
    elif role.startswith("recep"):
        keys = ["recep", "recept", "benef", "benefi", "destin"]
    else:
        # por si llega sin pista
        keys = ["recep","benef","emis","orden"]

    # todas las combinaciones razonables
    css_extra = []
    for k in keys:
        css_extra += [
            f'select[id*="{k}" i]', f'select[name*="{k}" i]',
            f'select[id*="cmb{k}" i]', f'select[name*="cmb{k}" i]',
        ]
    # ids más comunes exactos
    css_extra += [
        "#cmbInstitucionReceptora", "#cmbInstitucionEmisora",
        "#institucionReceptora", "#institucionEmisora",
        "#institucionBeneficiaria", "#institucionOrdenante",
    ]

    for css in css_extra:
        try:
            loc = frame.locator(css)
            if loc and loc.count():
                candidates.append(loc)
        except Exception:
            pass

    def _try_value_on(sel) -> bool:
        # (a) intento directo por label
        try:
            sel.select_option(label=wanted_text)
            # validar
            if _norm_txt(_get_selected_text(sel)) in ({desired} | aliases_norm):
                return True
        except Exception:
            pass

        # (b) barrido de <option> con comparación normalizada
        try:
            opts = sel.locator("option")
            cnt = opts.count()
            best_val, best_score = None, -1
            for i in range(cnt):
                o = opts.nth(i)
                txt = (o.inner_text() or "").strip()
                val = (o.get_attribute("value") or "").strip()
                if not (txt or val):
                    continue
                nt = _norm_txt(txt)
                score = 0
                if nt == desired or nt in aliases_norm:
                    score = 100
                elif nt.startswith(desired) or desired in nt:
                    score = 80
                if score > best_score:
                    best_score, best_val = score, (val or txt)
            if best_val is not None:
                sel.select_option(value=best_val)
                if _norm_txt(_get_selected_text(sel)) in ({desired} | aliases_norm):
                    return True
        except Exception:
            pass

        # (c) último recurso: JavaScript que ajusta selectedIndex + change
        try:
            ok = sel.evaluate(_JS_PICK_OPTION, wanted_text, list(aliases_norm))
            if ok and _norm_txt(_get_selected_text(sel)) in ({desired} | aliases_norm):
                return True
        except Exception:
            pass

        return False

    # probar candidatos
    for cand in candidates:
        if _try_value_on(cand):
            return

    # fallback extremo: cualquier <select> visible
    try:
        for sel in frame.locator("select:visible").all():
            if _try_value_on(sel):
                return
    except Exception:
        pass


def _set_date_in_picker(frame, date_str_ddmmyyyy: str) -> None:
    fi = frame.get_by_label(re.compile(r"Fecha.*realiz[oó]\s+el\s+pago", re.I))

    try:
        fi.click()
        try: fi.press("Control+A")
        except Exception: pass
        try: fi.press("Delete")
        except Exception: pass
        for ch in date_str_ddmmyyyy:
            fi.type(ch, delay=30)
        try:
            val = fi.input_value(timeout=800).strip()
            if val == date_str_ddmmyyyy:
                return
        except Exception:
            pass
    except Exception:
        pass

    try:
        fi.click()
        cont = _datepicker_container(frame)
        cont.wait_for(state="visible", timeout=3000)

        d, m, y = date_str_ddmmyyyy.split("-")
        target_day, target_month, target_year = int(d), int(m), int(y)

        try:
            cur_year, cur_month = frame.evaluate(
                "[new Date().getFullYear(), new Date().getMonth() + 1]"
            )
            cur_year, cur_month = int(cur_year), int(cur_month)
        except Exception:
            from datetime import datetime
            now = datetime.now()
            cur_year, cur_month = now.year, now.month

        prev_btn = cont.locator(
            "th.prev, .ui-datepicker-prev, button[aria-label*='Prev' i], button[aria-label*='Anterior' i]"
        ).first
        next_btn = cont.locator(
            "th.next, .ui-datepicker-next, button[aria-label*='Next' i], button[aria-label*='Siguiente' i]"
        ).first

        months_diff = (target_year - cur_year) * 12 + (target_month - cur_month)

        if months_diff != 0:
            btn = next_btn if months_diff > 0 else prev_btn
            for _ in range(abs(months_diff)):
                try:
                    btn.click(force=True)
                except Exception:
                    prev_btn = cont.locator(
                        "th.prev, .ui-datepicker-prev, button[aria-label*='Prev' i], button[aria-label*='Anterior' i]"
                    ).first
                    next_btn = cont.locator(
                        "th.next, .ui-datepicker-next, button[aria-label*='Next' i], button[aria-label*='Siguiente' i]"
                    ).first
                    btn = next_btn if months_diff > 0 else prev_btn
                    btn.click(force=True)
                frame.wait_for_timeout(60)

        day_css = f"td.day:not(.old):not(.new):has-text('{target_day}')"
        try:
            if cont.locator(day_css).count():
                cont.locator(day_css).first.click(force=True)
            else:
                cont.locator(
                    f"//td[normalize-space()='{target_day}' and not(contains(@class,'old')) and not(contains(@class,'new'))]"
                ).first.click(force=True)
        except Exception:
            pass

        try:
            val = fi.input_value(timeout=800).strip()
            if val == date_str_ddmmyyyy:
                return
        except Exception:
            pass

    except Exception:
        pass

    _force_set_input_value(fi, date_str_ddmmyyyy)

def _find_form_frame(page):
    try:
        targets = [
            re.compile(r"Fecha.*realiz[oó]\s+el\s+pago", re.I),
            re.compile(r"Criterio\s+de\s+b[uú]squeda", re.I),
            re.compile(r"Clave\s+de\s+rastreo|N[uú]mero\s+de\s+referencia", re.I),
        ]
        for fr in page.frames:
            ok = 0
            for rx in targets:
                try:
                    if fr.get_by_label(rx).first.count():
                        ok += 1
                except Exception:
                    pass
            if ok:
                return fr
    except Exception:
        pass
    return page.main_frame

# -------------------------------------------------------------------
# Automatización Playwright (descarga)
# -------------------------------------------------------------------
def _option_select_by_text(frame, select_locator: str, option_text: Optional[str]) -> None:
    if not option_text:
        return
    try:
        frame.locator(select_locator).select_option(label=option_text)
        return
    except Exception:
        pass
    try:
        for o in frame.locator(select_locator + " option").all():
            t = (o.inner_text() or "").strip().upper()
            if option_text.upper() in t:
                val = o.get_attribute("value") or ""
                frame.locator(select_locator).select_option(value=val)
                break
    except Exception:
        pass

def download_cep(job: CepJob, target_dir: Path, headless: bool = True, slowmo: int = 0) -> Optional[Path]:
    """
    Llena el formulario y descarga el CEP (PDF).
    Campos mínimos:
      - fecha dd-mm-YYYY
      - (clave de rastreo) o (número de referencia)
      - CLABE beneficiaria de 18 dígitos
      - monto > 0
    """
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

    target_dir.mkdir(parents=True, exist_ok=True)
    keyname = job.clave_rastreo or job.numero_referencia or "SINCLAVE"
    out_path = target_dir / f"{job.fecha}_{keyname[:24]}_{job.monto:.2f}.pdf"

    with sync_playwright() as p:
        # Argumentos para ejecutar en servidor sin GUI
        launch_args = {
            "headless": headless,
            "slow_mo": slowmo,
        }
        
        # Si es headless, agregar argumentos adicionales para evitar errores en servidores sin X
        if headless:
            launch_args["args"] = [
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu",
            ]
        
        browser = p.chromium.launch(**launch_args)
        ctx = browser.new_context(accept_downloads=True)
        page = ctx.new_page()
        page.set_default_timeout(35000)

        # AUTO-SAVE por si abre en popup o se hace click manual
        import time as _time
        saved_flag = {"done": False}

        def _auto_save(download):
            try:
                dest = out_path if not saved_flag["done"] else out_path.with_name(
                    f"{out_path.stem}__{int(_time.time())}.pdf"
                )
                download.save_as(str(dest))
                saved_flag["done"] = True
                print("[CEP] PDF guardado en:", dest)
            except Exception as e:
                print("[CEP] Error al auto-guardar:", e)

        ctx.on("download", _auto_save)

        try:
            page.goto("https://www.banxico.org.mx/cep/", wait_until="domcontentloaded")

            try:
                page.get_by_role("button", name=re.compile("Aceptar|De acuerdo|OK", re.I)).click(timeout=3000)
            except Exception:
                pass

            frm = _find_form_frame(page)

            _set_date_in_picker(frm, job.fecha)

            crit_select = 'select[name*="criterio"]'
            if job.clave_rastreo:
                _option_select_by_text(frm, crit_select, "Clave de rastreo")
                frm.get_by_label(re.compile(r"(Clave|N[uú]mero)\s+de\s+(rastreo|referencia)", re.I)).fill(job.clave_rastreo)
            else:
                _option_select_by_text(frm, crit_select, "Número de referencia")
                frm.get_by_label(re.compile(r"(Clave|N[uú]mero)\s+de\s+(rastreo|referencia)", re.I)).fill(job.numero_referencia or "")

            _select_option_by_label_loose(
                frm, re.compile(r"Instituci[oó]n\s+emisora\s+del\s+pago", re.I),
                job.banco_emisor, role_hint="emisora"
            )
            _select_option_by_label_loose(
                frm, re.compile(r"Instituci[oó]n\s+receptora\s+del\s+pago", re.I),
                job.banco_receptor, role_hint="receptora"
            )
            # Fallback directo por IDs nativos de Banxico (del HTML compartido)
            try:
                if job.banco_emisor:
                    sel = frm.locator("#input_emisor")
                    if sel and sel.count():
                        sel.select_option(label=job.banco_emisor)
            except Exception:
                pass

            try:
                if job.banco_receptor:
                    sel = frm.locator("#input_receptor")
                    if sel and sel.count():
                        sel.select_option(label=job.banco_receptor)
            except Exception:
                pass
        

            frm.wait_for_timeout(1200)

            if job.cuenta_beneficiaria:
                frm.get_by_label(re.compile(r"Cuenta\s+Beneficiaria", re.I)).fill(job.cuenta_beneficiaria)

            monto_input = frm.get_by_label(re.compile(r"Monto\s+del\s+pago", re.I))
            monto_input.fill(f"{float(job.monto):.2f}")
            try:
                monto_input.evaluate("""
                    el => {
                        el.dispatchEvent(new Event('input',{bubbles:true}));
                        el.dispatchEvent(new Event('change',{bubbles:true}));
                        if (el.blur) el.blur();
                    }
                """)
            except Exception:
                pass

            frm.wait_for_timeout(1200)

            if not _handle_captcha(frm, page, out_path.parent, headless=headless):
                print("[CEP] Captcha no resuelto (headless). Saltando este movimiento.")
                return None

            pth = _click_descargar_y_bajar_pdf(page, frm, out_path)
            return pth

        except PWTimeout:
            return None
        except Exception as e:
            print("[CEP] Excepción en download_cep:", repr(e))
            return None
        finally:
            try:
                if not headless and os.getenv("CEP_DEBUG_KEEP", "0") == "1":
                    page.wait_for_timeout(8000)
            except Exception:
                pass
            ctx.close()
            browser.close()

# -------------------------------------------------------------------
# Escribir hipervínculos y generar ZIP (con CSV de log)
# -------------------------------------------------------------------
def _add_links_to_excel(xlsx_in: str, xlsx_out: str, jobs: List[CepJob], pdf_paths: Dict[Tuple[str, int], Path]) -> None:
    wb = load_workbook(xlsx_in)
    for sheet_name in list(wb.sheetnames):
        ws = wb[sheet_name]
        cep_col_idx = ws.max_column + 1
        ws.cell(row=1, column=cep_col_idx, value="CEP")
        for j in jobs:
            if j.sheet != sheet_name:
                continue
            pdf = pdf_paths.get((j.sheet, j.row_index))
            if not pdf:
                continue
            excel_row = j.row_index + 2
            rel = Path("ceps") / j.sheet / pdf.name
            c = ws.cell(row=excel_row, column=cep_col_idx)
            c.value = "Abrir CEP"
            c.hyperlink = str(rel)
            c.style = "Hyperlink"
        ws.column_dimensions[get_column_letter(cep_col_idx)].width = 14
    wb.save(xlsx_out)

def _zip_package(zip_path: str, excel_path: str, ceps_root: Path) -> None:
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        z.write(excel_path, arcname=Path(excel_path).name)
        if ceps_root.exists():
            for p in ceps_root.rglob("*.pdf"):
                arc = Path("ceps") / p.relative_to(ceps_root)
                z.write(p, arcname=str(arc))
            log_csv = ceps_root / "_descargas_cep.csv"
            if log_csv.exists():
                z.write(log_csv, arcname=str(Path("ceps") / "_descargas_cep.csv"))

# -------------------------------------------------------------------
# Orquestadores (desde Excel o desde PDF)
# -------------------------------------------------------------------
def build_zip_with_ceps_from_xlsx(xlsx_path: str, zip_out: str, headless: bool = True) -> Dict:
    tmp_dir = Path(".cep_work")
    ceps_root = tmp_dir / "ceps"
    ceps_root.mkdir(parents=True, exist_ok=True)

    jobs, _ = collect_jobs_from_xlsx(xlsx_path)

    log_csv = ceps_root / "_descargas_cep.csv"
    import csv
    with log_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sheet","row_index","fecha","monto","clave_rastreo","referencia","status","detalle"])

        pdf_by_row: Dict[Tuple[str, int], Path] = {}
        ok = 0

        if jobs:
            try:
                from playwright.sync_api import sync_playwright  # noqa
            except Exception:
                raise RuntimeError(
                    "Playwright no está instalado.\n"
                    "  pip install playwright\n"
                    "  playwright install --with-deps chromium"
                )

            for j in jobs:
                clabe = (j.cuenta_beneficiaria or "").strip()
                if not re.fullmatch(r"\d{18}", clabe):
                    w.writerow([j.sheet, j.row_index, j.fecha, j.monto, j.clave_rastreo or "", j.numero_referencia or "", "skip", "missing_clabe_18d"])
                    continue
                if not (j.clave_rastreo or j.numero_referencia):
                    w.writerow([j.sheet, j.row_index, j.fecha, j.monto, "", "", "skip", "missing_key"])
                    continue
                if not j.monto or float(j.monto) <= 0:
                    w.writerow([j.sheet, j.row_index, j.fecha, j.monto, j.clave_rastreo or "", j.numero_referencia or "", "skip", "missing_amount"])
                    continue

                target = ceps_root / j.sheet
                target.mkdir(parents=True, exist_ok=True)
                try:
                    pth = download_cep(j, target, headless=headless)
                    if pth and pth.exists():
                        ok += 1
                        pdf_by_row[(j.sheet, j.row_index)] = pth
                        w.writerow([j.sheet, j.row_index, j.fecha, j.monto, j.clave_rastreo or "", j.numero_referencia or "", "ok", pth.name])
                    else:
                        w.writerow([j.sheet, j.row_index, j.fecha, j.monto, j.clave_rastreo or "", j.numero_referencia or "", "fail", "button_disabled_or_timeout"])
                except Exception as e:
                    w.writerow([j.sheet, j.row_index, j.fecha, j.monto, j.clave_rastreo or "", j.numero_referencia or "", "fail", repr(e)])
        else:
            pdf_by_row = {}
            ok = 0

    excel_out = tmp_dir / (Path(xlsx_path).stem + "_con_ceps.xlsx")
    _add_links_to_excel(xlsx_path, str(excel_out), jobs, pdf_by_row)
    _zip_package(zip_out, str(excel_out), ceps_root)

    shutil.rmtree(tmp_dir, ignore_errors=True)
    return {"total_jobs": len(jobs), "ok": sum(1 for _ in (pdf_by_row.keys())), "fail": len(jobs) - sum(1 for _ in (pdf_by_row.keys())), "zip": zip_out}

def build_zip_with_ceps_from_pdf(pdf_path: str, extractor: str, zip_out: str, headless: bool = True) -> Dict:
    tmp_xlsx = Path(".cep_work") / (Path(pdf_path).stem + ".xlsx")
    tmp_xlsx.parent.mkdir(parents=True, exist_ok=True)

    ext = (extractor or "").strip().lower()
    if ext == "santander" and extract_santander_to_xlsx:
        extract_santander_to_xlsx(pdf_path, str(tmp_xlsx))
    elif ext == "bbva" and extract_bbva_to_xlsx:
        extract_bbva_to_xlsx(pdf_path, str(tmp_xlsx))
    elif ext == "banorte" and extract_banorte_to_xlsx:
        extract_banorte_to_xlsx(pdf_path, str(tmp_xlsx))
    elif extractor.lower() == "inbursa":
        if not extract_inbursa_to_xlsx:
            raise RuntimeError("Extractor Inbursa no disponible")
        extract_inbursa_to_xlsx(pdf_path, str(tmp_xlsx))

    else:
        raise ValueError("Extractor desconocido o no disponible. Usa: santander | bbva | banorte")

    try:
        return build_zip_with_ceps_from_xlsx(str(tmp_xlsx), zip_out, headless=headless)
    finally:
        shutil.rmtree(tmp_xlsx.parent, ignore_errors=True)

# -------------------------------------------------------------------
# CLI
# -------------------------------------------------------------------
def _parse_args(argv: List[str]) -> Dict:
    import argparse
    p = argparse.ArgumentParser(description="Descarga CEPs de Banxico y genera ZIP con Excel + PDFs.")
    p.add_argument("--xlsx", help="Excel de movimientos ya generado por el extractor.")
    p.add_argument("--pdf", help="PDF del estado de cuenta (si quieres que este script llame al extractor).")
    p.add_argument("--extractor", choices=["santander","bbva","banorte"], help="Extractor a usar cuando se pasa --pdf.")
    p.add_argument("--zip", required=True, help="Ruta de salida del ZIP.")
    p.add_argument("--headed", action="store_true", help="Mostrar el navegador (debug).")
    args = p.parse_args(argv)
    if not args.xlsx and not args.pdf:
        p.error("Debes indicar --xlsx o --pdf.")
    if args.pdf and not args.extractor:
        p.error("Si usas --pdf debes indicar --extractor.")
    return {
        "xlsx": args.xlsx, "pdf": args.pdf, "extractor": args.extractor,
        "zip": args.zip, "headless": not args.headed
    }

def main(argv: List[str]) -> None:
    cfg = _parse_args(argv)
    if cfg["xlsx"]:
        res = build_zip_with_ceps_from_xlsx(cfg["xlsx"], cfg["zip"], headless=cfg["headless"])
    else:
        res = build_zip_with_ceps_from_pdf(cfg["pdf"], cfg["extractor"], cfg["zip"], headless=cfg["headless"])
    print(json.dumps(res, indent=2, ensure_ascii=False))

if __name__ == "__main__":
    main(sys.argv[1:])
