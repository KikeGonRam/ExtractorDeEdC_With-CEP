# app.py
from __future__ import annotations
import jwt

from dotenv import load_dotenv
load_dotenv()
from unicodedata import normalize as _u_norm
import re

import sys, asyncio
import os, shutil, tempfile, csv, io, hashlib, logging
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, Optional

from fastapi import FastAPI, UploadFile, File, BackgroundTasks, HTTPException, Form, Depends, Request
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi import HTTPException
from fastapi.middleware.cors import CORSMiddleware
from starlette.concurrency import run_in_threadpool

logger = logging.getLogger("uvicorn.error")
if not logger.handlers:
    logging.basicConfig(level=logging.INFO)

# --- Windows: loop con soporte de subprocess
if sys.platform.startswith("win"):
    try:
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    except Exception:
        pass


JWT_SECRET = os.getenv("JWT_SECRET", "supersecreto123")

def verify_jwt(request: Request):
    # Prefer standard Authorization header but allow legacy custom header
    token = request.headers.get("auth_token") or request.headers.get("Authorization")
    if token and isinstance(token, str) and token.startswith("Bearer "):
        token = token.split(" ", 1)[1]
    if not token:
        # Log header keys for debugging (do NOT log header values / token)
        try:
            hdrs = list(request.headers.keys())
        except Exception:
            hdrs = []
        logger.warning("JWT requerido pero no se encontró token en headers. Header keys: %s", hdrs)
        raise HTTPException(status_code=401, detail="Token JWT requerido")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        return payload  # Puedes retornar el usuario si lo necesitas
    except Exception as e:
        # Evita lanzar AttributeError si la librería jwt difiere en excepciones
        try:
            hdrs = list(request.headers.keys())
        except Exception:
            hdrs = []
        logger.warning("Falló la verificación JWT (%s). Header keys: %s", type(e).__name__, hdrs)
        raise HTTPException(status_code=403, detail="Token JWT inválido")

# ==========================
#   Rutas de estáticos y store
# ==========================
BASE_DIR = Path(__file__).parent.resolve()
STATIC_DIR = BASE_DIR / "static"
STATIC_DIR.mkdir(exist_ok=True)

STORE_DIR = BASE_DIR / "store"
INPUT_DIR = STORE_DIR / "inputs"
OUTPUT_DIR = STORE_DIR / "outputs"
INPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ==========================
#   Extractores
# ==========================
import santander_extractor as santander

EXTRACT_MAP: Dict[str, Callable[[str, str], None]] = {
    "santander": santander.extract_santander_to_xlsx,
}

def _get_extractor(bank: str) -> Callable[[str, str], None]:
    b = (bank or "").strip().lower()
    if b not in EXTRACT_MAP:
        if b == "banorte":
            from banorte_extractor import extract_banorte_to_xlsx
            EXTRACT_MAP[b] = extract_banorte_to_xlsx
        elif b == "bbva":
            from bbva_extractor import extract_bbva_to_xlsx
            EXTRACT_MAP[b] = extract_bbva_to_xlsx
        elif b == "inbursa":
            from inbursa_extractor import extract_inbursa_to_xlsx
            EXTRACT_MAP[b] = extract_inbursa_to_xlsx
        else:
            raise HTTPException(status_code=400, detail=f"Banco no soportado: {bank}")
    return EXTRACT_MAP[b]

def _get_cep_service():
    try:
        from cep_service import make_zip_with_ceps_for_bank
        return make_zip_with_ceps_for_bank
    except Exception:
        return None

def _call_ceps(make_zip_fn, bank: str, pdf_path: str, workdir: str) -> str:
    """acepta firmas input_pdf/pdf_path/input_path indistintamente."""
    try:
        return make_zip_fn(bank=bank, input_pdf=pdf_path, workdir=workdir)
    except TypeError:
        try:
            return make_zip_fn(bank=bank, pdf_path=pdf_path, workdir=workdir)
        except TypeError:
            return make_zip_fn(bank=bank, input_path=pdf_path, workdir=workdir)

# ==========================
#   Helpers
# ==========================
def _mk_workdir() -> Path:
    return Path(tempfile.mkdtemp(prefix="extract_"))

def _schedule_cleanup(background_tasks: BackgroundTasks, *paths: Path):
    for p in paths:
        if p.is_dir():
            background_tasks.add_task(shutil.rmtree, str(p), True)
        elif p.is_file():
            background_tasks.add_task(os.remove, str(p))

def _copy_for_response(src: Path) -> Path:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=src.suffix)
    tmp_path = Path(tmp.name); tmp.close()
    shutil.copy2(src, tmp_path)
    return tmp_path

def _as_error(e: Exception, status: int = 500):
    logger.exception("Fallo en request")
    return JSONResponse({"error": f"{type(e).__name__}: {e}"}, status_code=status)

def _sha256_of_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def _save_pdf_to_store(source_path: Path) -> tuple[Optional[int], Optional[str]]:
    """Guarda el PDF de entrada en store/inputs/{sha}.pdf"""
    try:
        tam = source_path.stat().st_size
    except Exception:
        tam = None
    try:
        sha = _sha256_of_file(source_path)
        if sha:
            target_pdf = INPUT_DIR / f"{sha}.pdf"
            if not target_pdf.exists():
                shutil.copy2(source_path, target_pdf)
    except Exception:
        sha = None
        logger.exception("No se pudo calcular/guardar SHA del PDF")
    return tam, sha

def _save_output_to_store(out_path: Path, ext: str) -> tuple[Optional[int], Optional[str], Path]:
    """
    Guarda XLSX/ZIP en store/outputs/{sha}.{ext} y devuelve
    (tamaño, sha, ruta_en_store).
    """
    try:
        tam = out_path.stat().st_size
    except Exception:
        tam = None
    try:
        sha = _sha256_of_file(out_path)
        if not sha:
            raise ValueError("SHA vacío")
        dst = OUTPUT_DIR / f"{sha}.{ext}"
        if not dst.exists():
            shutil.copy2(out_path, dst)
        return tam, sha, dst
    except Exception:
        logger.exception("No se pudo guardar salida en store")
        return tam, None, out_path


# ==== LECTURA DE EMPRESA DESDE XLSX ====
try:
    from openpyxl import load_workbook
    _HAVE_OPENPYXL = True
except Exception:
    _HAVE_OPENPYXL = False
    logger.warning("openpyxl no disponible: no se leerá 'empresa' desde XLSX.")

# ==== LECTURA DE EMPRESA DESDE XLSX (mejorada) ====

try:
    from openpyxl import load_workbook
    _HAVE_OPENPYXL = True
except Exception:
    _HAVE_OPENPYXL = False
    logger.warning("openpyxl no disponible: no se leerá 'empresa' desde XLSX.")

def _norm_key(s: str) -> str:
    """Normaliza para comparar: quita acentos, deja alfanum, espacios simples y minúsculas."""
    if not isinstance(s, str):
        s = str(s) if s is not None else ""
    s = _u_norm("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9 ]+", " ", s).lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s

_HEADER_TOKENS = {
    "empresa", "razon social", "rfc", "fecha", "banco", "periodo", "moneda",
    "cuenta", "archivo", "folio", "sucursal", "cliente", "titular",
    "reporte", "mes", "anio", "ano", "año", "correo", "email", "saldo inicial", "saldo final",
}
_TARGET_LABELS = {"empresa", "razon social"}  # etiquetas que buscamos

def _is_header_like(s: str) -> bool:
    k = _norm_key(s)
    return (k in _HEADER_TOKENS) or (len(k) <= 2)

def _first_non_header(values) -> Optional[str]:
    for v in values:
        if v is None:
            continue
        if not isinstance(v, str):
            v = str(v)
        v = v.strip()
        if not v:
            continue
        if not _is_header_like(v):
            return v[:255]
    return None

def _extract_empresa_from_xlsx(xlsx_path: Path) -> Optional[str]:
    """
    Busca el nombre de la empresa en la hoja 'Info' (o similares).
    Preferimos SIEMPRE el valor **debajo** de la celda 'Empresa'/'Razón Social'.
    Si eso falla, intentamos a la derecha (saltando encabezados) y casos tipo 'Empresa: Valor'.
    """
    if not _HAVE_OPENPYXL or not xlsx_path.exists():
        return None

    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
    except Exception:
        logger.exception("No se pudo abrir XLSX para extraer 'empresa'")
        return None

    # Elegimos hoja: preferimos 'info*'
    ws = None
    for s in wb.worksheets:
        t = _norm_key(s.title)
        if t == "info" or t.startswith("info "):
            ws = s
            break
    if ws is None:
        ws = wb.worksheets[0]

    max_r = min(ws.max_row or 0, 60) or 60
    max_c = min(ws.max_column or 0, 12) or 12

    # 1) Buscamos la celda etiqueta y tomamos valores DEBAJO (preferido) y DERECHA como respaldo.
    for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            key = _norm_key(val)
            if key in _TARGET_LABELS:
                r, c = cell.row, cell.column

                # Preferimos "debajo": primeras 4 celdas bajo la etiqueta
                down_vals = [ws.cell(row=r + i, column=c).value for i in range(1, 5)]
                # Respaldo: a la derecha, primeras 4 celdas (para diseños horizontales)
                right_vals = [ws.cell(row=r, column=c + i).value for i in range(1, 5)]

                cand = _first_non_header(down_vals) or _first_non_header(right_vals)
                if cand:
                    return cand

    # 2) Casos "Empresa: ACME S.A."
    for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and ":" in v:
                left, right = v.split(":", 1)
                if _norm_key(left) in _TARGET_LABELS:
                    right = right.strip()
                    if right and not _is_header_like(right):
                        return right[:255]

    return None

def _update_empresa(row_id: int, empresa: Optional[str]) -> None:
    """Actualiza el campo empresa si hay valor."""
    if engine is None or not row_id or not empresa:
        return
    safe = empresa.strip()
    if not safe:
        return
    try:
        with engine.begin() as conn:
            conn.execute(
                solicitudes.update()
                .where(solicitudes.c.id == row_id)
                .values(empresa=safe[:255])
            )
    except Exception:
        logger.exception("No se pudo actualizar 'empresa' en BD")

# =========================================================
#   SQLAlchemy
# =========================================================
from sqlalchemy import (
    create_engine, MetaData, Table, Column, Integer, BigInteger,
    String, Text, DateTime, func, select, and_, desc
)
from sqlalchemy.engine import URL

def _db_url_from_env() -> Optional[URL]:
    host = os.getenv("DB_HOST"); port = os.getenv("DB_PORT", "3306")
    name = os.getenv("DB_NAME"); user = os.getenv("DB_USER"); pwd = os.getenv("DB_PASS")
    if not (host and name and user and pwd):
        return None
    try:
        return URL.create(
            "mysql+pymysql",
            username=user, password=pwd,
            host=host, port=int(port), database=name,
            query={"charset": "utf8mb4"},
        )
    except Exception:
        logger.exception("Error armando URL de BD")
        return None

DB_URL = _db_url_from_env()
metadata = MetaData()

# Esquema (incluye columnas de salida)
solicitudes = Table(
    "solicitudes", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("id_usuario", Integer, nullable=False),
    Column("nombre_usuario", String(255), nullable=False),
    Column("archivo_nombre", String(255), nullable=False),
    Column("archivo_tamano", BigInteger, nullable=True),
    Column("archivo_sha256", String(64), nullable=True),
    Column("salida_nombre", String(255), nullable=True),
    Column("salida_tamano", BigInteger, nullable=True),
    Column("salida_sha256", String(64), nullable=True),
    Column("banco", String(32), nullable=False),          # enum en BD; aquí como String
    Column("empresa", String(255), nullable=False),
    Column("solicitado_en", DateTime, server_default=func.current_timestamp(), nullable=False),
    Column("resultado", String(16), nullable=False),       # 'xlsx' | 'zip'
    Column("estado", String(16), nullable=False),          # 'ok' | 'fail' | 'processing'
    Column("error", Text, nullable=True),
)

engine = create_engine(DB_URL, pool_pre_ping=True, pool_recycle=280, echo=False) if DB_URL else None
if not engine:
    logger.warning("DB_URL no configurado. No se registrarán solicitudes en BD.")

def _init_db():
    if engine is None: return
    try:
        metadata.create_all(engine)
    except Exception:
        logger.exception("No se pudo crear/verificar la tabla 'solicitudes'.")

# ==========================
#   Mapeos a ENUM de BD
# ==========================
RESULT_APP2DB = {"excel": "xlsx", "excel+cep": "zip"}
RESULT_DB2APP = {"xlsx": "excel", "zip": "excel+cep"}

EST_OK, EST_FAIL, EST_PROC = "ok", "fail", "processing"

# ==========================
#   Registro de solicitudes (begin/finish)
# ==========================
def _insert_processing(
    archivo_nombre: str, banco: str, empresa: Optional[str],
    resultado_db: str, tam: Optional[int], sha: Optional[str],
    id_usuario: int, nombre_usuario: str
) -> int:
    if engine is None:
        return 0
    safe_empresa = (empresa or "").strip() or "SIN_EMPRESA"
    with engine.begin() as conn:
        res = conn.execute(
            solicitudes.insert().values(
                id_usuario=id_usuario,
                nombre_usuario=nombre_usuario,
                archivo_nombre=(archivo_nombre or "archivo.pdf")[:255],
                archivo_tamano=tam,
                archivo_sha256=sha,
                banco=(banco or "").strip().lower()[:32],
                empresa=safe_empresa[:255],
                resultado=resultado_db,
                estado=EST_PROC,   # En proceso
                error=None,
            )
        )
        return int(res.inserted_primary_key[0])

def _update_finish(
    row_id: int, estado: str, error: Optional[str] = None,
    *, salida_nombre: Optional[str] = None,
    salida_tamano: Optional[int] = None,
    salida_sha256: Optional[str] = None,
):
    if engine is None or not row_id:
        return
    with engine.begin() as conn:
        conn.execute(
            solicitudes.update().where(solicitudes.c.id == row_id).values(
                estado=estado,
                error=(error[:2000] if error else None),
                salida_nombre=salida_nombre[:255] if salida_nombre else None,
                salida_tamano=salida_tamano,
                salida_sha256=salida_sha256,
            )
        )

# ==========================
#   FastAPI app + CORS
# ==========================
app = FastAPI(title="Extractor de estados (Banorte / BBVA / Santander / Inbursa)")
# lee orígenes desde env o usa los típicos de dev
_CORS = os.getenv("CORS_ORIGINS", "http://localhost:3000,http://127.0.0.1:3000").split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in _CORS if o.strip()],
    allow_credentials=False,          # no usamos cookies, evita problemas con '*'
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

@app.on_event("startup")
def _on_start():
    _init_db()
    try:
        STORE_DIR.mkdir(exist_ok=True)
        INPUT_DIR.mkdir(parents=True, exist_ok=True)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        logger.exception("No se pudo preparar carpetas del store")

# ==========================
#   Rutas básicas
# ==========================
@app.get("/")
def home():
    index_path = STATIC_DIR / "index.html"
    if not index_path.exists():
        return JSONResponse({"detail": f"No se encontró {index_path}. Coloca tu index.html dentro de /static junto a app.py."}, status_code=404)
    return FileResponse(str(index_path))

@app.get("/healthz")
def healthz():
    return {"ok": True, "static_dir": str(STATIC_DIR), "has_index": (STATIC_DIR / "index.html").exists()}

@app.get("/solicitudes/stats")
def solicitudes_stats(
    banco: str | None = None,
    resultado: str | None = None,   # 'xlsx' | 'zip'
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    q: str | None = None,
    empresa: str | None = None,
):
    if engine is None:
        raise HTTPException(status_code=503, detail="BD no configurada")

    d1 = _parse_date_yyyy_mm_dd(fecha_desde)
    d2 = _parse_date_yyyy_mm_dd(fecha_hasta)

    conds = []
    if banco: conds.append(solicitudes.c.banco == banco.strip().lower())
    if resultado in {"xlsx", "zip"}: conds.append(solicitudes.c.resultado == resultado)
    if d1: conds.append(solicitudes.c.solicitado_en >= d1)
    if d2: conds.append(solicitudes.c.solicitado_en < d2.replace(hour=23, minute=59, second=59, microsecond=999999))
    if empresa: conds.append(solicitudes.c.empresa.ilike(f"%{empresa.strip()}%"))
    if q:
        like = f"%{q.strip()}%"
        conds.append((solicitudes.c.archivo_nombre.ilike(like)) | (solicitudes.c.empresa.ilike(like)))

    where_all = and_(*conds) if conds else None

    with engine.begin() as conn:
        total = conn.execute(
            select(func.count()).select_from(solicitudes).where(where_all) if where_all
            else select(func.count()).select_from(solicitudes)
        ).scalar_one()

        rows = conn.execute(
            select(solicitudes.c.estado, func.count().label("c"))
            .where(where_all) if where_all else
            select(solicitudes.c.estado, func.count().label("c"))
        ).group_by(solicitudes.c.estado).all()

    counts = {"ok": 0, "processing": 0, "fail": 0}
    for r in rows:
        counts[str(r[0])] = int(r[1])
    return {"total": int(total), **counts}


@app.get("/file/{solicitud_id}")
def open_original_file(solicitud_id: int):
    if engine is None:
        raise HTTPException(status_code=503, detail="BD no configurada")

    with engine.begin() as conn:
        row = conn.execute(
            select(solicitudes.c.archivo_nombre, solicitudes.c.archivo_sha256)
            .where(solicitudes.c.id == solicitud_id)
        ).mappings().first()

    if not row or not row["archivo_sha256"]:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    src = INPUT_DIR / f"{row['archivo_sha256']}.pdf"
    if not src.exists():
        raise HTTPException(status_code=404, detail="PDF no disponible en el repositorio")

    # devuelve con un nombre amigable
    return FileResponse(
        path=str(src),
        filename=row["archivo_nombre"] or f"archivo_{solicitud_id}.pdf",
        media_type="application/pdf",
    )

# =========================================================
#   RUTAS GENÉRICAS
# =========================================================
@app.post("/extract/{bank}")
async def extract_generic(
    bank: str, background_tasks: BackgroundTasks,
    file: UploadFile = File(...), 
    empresa: Optional[str] = Form(None),
    user: dict = Depends(verify_jwt)
):
    original_name = file.filename or "estado.pdf"
    pdf_path: Optional[Path] = None
    row_id = 0
    try:
        extractor = _get_extractor(bank)
        workdir = _mk_workdir()
        pdf_path = workdir / original_name
        pdf_path.write_bytes(await file.read())

        tam, sha = _save_pdf_to_store(pdf_path)
        id_usuario = user.get("id_usuario")
        nombre_usuario = user.get("nombre")
        row_id = _insert_processing(original_name, bank, empresa, RESULT_APP2DB["excel"], tam, sha, id_usuario, nombre_usuario)

        out_xlsx = workdir / (Path(original_name).stem + ".xlsx")
        await run_in_threadpool(extractor, str(pdf_path), str(out_xlsx))

        # === Empresa desde XLSX ===
        try:
            emp = _extract_empresa_from_xlsx(out_xlsx)
            if emp:
                _update_empresa(row_id, emp)
        except Exception:
            logger.exception("No se pudo extraer/actualizar 'empresa' desde XLSX")

        # Guardar salida en store/outputs y registrar en BD
        tam_out, sha_out, stored = _save_output_to_store(out_xlsx, "xlsx")
        _update_finish(row_id, EST_OK, salida_nombre=out_xlsx.name, salida_tamano=tam_out, salida_sha256=sha_out)

        tmp_path = _copy_for_response(stored)
        _schedule_cleanup(background_tasks, workdir, tmp_path)

        return FileResponse(path=tmp_path, filename=out_xlsx.name,
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        _update_finish(row_id, EST_FAIL, str(e))
        return _as_error(e)

@app.post("/extract-with-cep/{bank}")
async def extract_with_cep_generic(
    bank: str, background_tasks: BackgroundTasks,
    file: UploadFile = File(...), empresa: Optional[str] = Form(None),
    user: dict = Depends(verify_jwt)
):
    original_name = file.filename or "estado.pdf"
    pdf_path: Optional[Path] = None
    row_id = 0
    try:
        make_zip_with_ceps_for_bank = _get_cep_service()
        if not make_zip_with_ceps_for_bank:
            # Si falta el servicio de CEP o Playwright no está instalado, devolver 501
            # en lugar de lanzar una excepción que produce un 500 genérico.
            return _as_error(Exception("Falta cep_service.py o Playwright. Instala playwright y crea cep_service.py."), status=501)

        workdir = _mk_workdir()
        pdf_path = workdir / original_name
        pdf_path.write_bytes(await file.read())

        tam, sha = _save_pdf_to_store(pdf_path)
        row_id = _insert_processing(original_name, bank, empresa, RESULT_APP2DB["excel+cep"], tam, sha)

        # === Generar XLSX temporal para leer empresa ===
        try:
            extractor_tmp = _get_extractor(bank)
            tmp_xlsx = Path(workdir) / (Path(original_name).stem + ".xlsx")
            await run_in_threadpool(extractor_tmp, str(pdf_path), str(tmp_xlsx))
            emp = _extract_empresa_from_xlsx(tmp_xlsx)
            if emp:
                _update_empresa(row_id, emp)
            try:
                if tmp_xlsx.exists():
                    tmp_xlsx.unlink()
            except Exception:
                pass
        except Exception:
            logger.exception("No se pudo generar XLSX temporal para leer 'empresa'")

        logger.info(f"[CEP DEBUG] Generando ZIP con CEP: bank={bank}, pdf_path={pdf_path}, workdir={workdir}")
        zip_str = await run_in_threadpool(_call_ceps, make_zip_with_ceps_for_bank, bank, str(pdf_path), str(workdir))
        logger.info(f"[CEP DEBUG] Resultado zip_str={zip_str}")
        zip_path = Path(zip_str)
        logger.info(f"[CEP DEBUG] zip_path.exists={zip_path.exists()}, zip_path={zip_path}")

        # Guardar salida en store/outputs y registrar en BD
        tam_out, sha_out, stored = _save_output_to_store(zip_path, "zip")
        _update_finish(
            row_id, EST_OK,
            salida_nombre=f"{Path(original_name).stem}_ceps.zip",
            salida_tamano=tam_out, salida_sha256=sha_out
        )

        tmp_path = _copy_for_response(stored)
        _schedule_cleanup(background_tasks, workdir, tmp_path)

        return FileResponse(path=tmp_path, filename=zip_path.name, media_type="application/zip")
    except Exception as e:
        _update_finish(row_id, EST_FAIL, str(e))
        return _as_error(e)

# =========================================================
#   ESPECÍFICAS (compatibilidad con frontend actual)
# =========================================================
@app.post("/extract/santander")
async def extract_santander_only(background_tasks: BackgroundTasks, file: UploadFile = File(...), empresa: Optional[str] = Form(None), user: dict = Depends(verify_jwt)):
    bank = "santander"
    original_name = file.filename or "estado.pdf"
    pdf_path: Optional[Path] = None
    row_id = 0
    try:
        workdir = _mk_workdir()
        pdf_path = workdir / original_name
        pdf_path.write_bytes(await file.read())
        tam, sha = _save_pdf_to_store(pdf_path)
        row_id = _insert_processing(original_name, bank, empresa, RESULT_APP2DB["excel"], tam, sha)

        out_xlsx = workdir / (Path(original_name).stem + "_santander.xlsx")
        await run_in_threadpool(santander.extract_santander_to_xlsx, str(pdf_path), str(out_xlsx))

        # === Empresa desde XLSX ===
        try:
            emp = _extract_empresa_from_xlsx(out_xlsx)
            if emp:
                _update_empresa(row_id, emp)
        except Exception:
            logger.exception("No se pudo extraer/actualizar 'empresa' desde XLSX")

        tam_out, sha_out, stored = _save_output_to_store(out_xlsx, "xlsx")
        _update_finish(row_id, EST_OK, salida_nombre=out_xlsx.name, salida_tamano=tam_out, salida_sha256=sha_out)

        tmp_path = _copy_for_response(stored)
        _schedule_cleanup(background_tasks, workdir, tmp_path)

        return FileResponse(path=tmp_path, filename=out_xlsx.name,
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        _update_finish(row_id, EST_FAIL, str(e))
        return _as_error(e)

@app.post("/extract/santander/with-cep")
async def extract_santander_with_cep(background_tasks: BackgroundTasks, file: UploadFile = File(...), empresa: Optional[str] = Form(None), user: dict = Depends(verify_jwt)):
    bank = "santander"
    original_name = file.filename or "estado.pdf"
    pdf_path: Optional[Path] = None
    row_id = 0
    try:
        make_zip_with_ceps_for_bank = _get_cep_service()
        if not make_zip_with_ceps_for_bank:
            return _as_error(Exception("Falta cep_service.py o Playwright."), status=501)

        workdir = _mk_workdir()
        pdf_path = workdir / original_name
        pdf_path.write_bytes(await file.read())
        tam, sha = _save_pdf_to_store(pdf_path)
        row_id = _insert_processing(original_name, bank, empresa, RESULT_APP2DB["excel+cep"], tam, sha)

        # === XLSX temporal para empresa ===
        try:
            extractor_tmp = _get_extractor(bank)
            tmp_xlsx = Path(workdir) / (Path(original_name).stem + ".xlsx")
            await run_in_threadpool(extractor_tmp, str(pdf_path), str(tmp_xlsx))
            emp = _extract_empresa_from_xlsx(tmp_xlsx)
            if emp:
                _update_empresa(row_id, emp)
            try:
                if tmp_xlsx.exists():
                    tmp_xlsx.unlink()
            except Exception:
                pass
        except Exception:
            logger.exception("No se pudo generar XLSX temporal para leer 'empresa'")

        zip_str = await run_in_threadpool(_call_ceps, make_zip_with_ceps_for_bank, bank, str(pdf_path), str(workdir))
        zip_path = Path(zip_str)

        tam_out, sha_out, stored = _save_output_to_store(zip_path, "zip")
        _update_finish(row_id, EST_OK, salida_nombre=zip_path.name, salida_tamano=tam_out, salida_sha256=sha_out)

        tmp_path = _copy_for_response(stored)
        _schedule_cleanup(background_tasks, workdir, tmp_path)

        return FileResponse(path=tmp_path, filename=zip_path.name, media_type="application/zip")
    except Exception as e:
        _update_finish(row_id, EST_FAIL, str(e))
        return _as_error(e)

@app.post("/extract/banorte")
async def extract_banorte_only(background_tasks: BackgroundTasks, file: UploadFile = File(...), empresa: Optional[str] = Form(None), user: dict = Depends(verify_jwt)):
    from banorte_extractor import extract_banorte_to_xlsx
    bank = "banorte"
    original_name = file.filename or "estado.pdf"
    pdf_path: Optional[Path] = None
    row_id = 0
    try:
        workdir = _mk_workdir()
        pdf_path = workdir / original_name
        pdf_path.write_bytes(await file.read())
        tam, sha = _save_pdf_to_store(pdf_path)
        row_id = _insert_processing(original_name, bank, empresa, RESULT_APP2DB["excel"], tam, sha)

        out_xlsx = workdir / (Path(original_name).stem + ".xlsx")
        await run_in_threadpool(extract_banorte_to_xlsx, str(pdf_path), str(out_xlsx))

        # === Empresa desde XLSX ===
        try:
            emp = _extract_empresa_from_xlsx(out_xlsx)
            if emp:
                _update_empresa(row_id, emp)
        except Exception:
            logger.exception("No se pudo extraer/actualizar 'empresa' desde XLSX")

        tam_out, sha_out, stored = _save_output_to_store(out_xlsx, "xlsx")
        _update_finish(row_id, EST_OK, salida_nombre=out_xlsx.name, salida_tamano=tam_out, salida_sha256=sha_out)

        tmp_path = _copy_for_response(stored); _schedule_cleanup(background_tasks, workdir, tmp_path)
        return FileResponse(path=tmp_path, filename=out_xlsx.name,
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        _update_finish(row_id, EST_FAIL, str(e))
        return _as_error(e)

@app.post("/extract/banorte/with-cep")
async def extract_banorte_with_cep(background_tasks: BackgroundTasks, file: UploadFile = File(...), empresa: Optional[str] = Form(None), user: dict = Depends(verify_jwt)):
    bank = "banorte"
    original_name = file.filename or "estado.pdf"
    pdf_path: Optional[Path] = None
    row_id = 0
    try:
        make_zip_with_ceps_for_bank = _get_cep_service()
        if not make_zip_with_ceps_for_bank:
            return _as_error(Exception("Falta cep_service.py o Playwright."), status=501)

        workdir = _mk_workdir()
        pdf_path = workdir / original_name
        pdf_path.write_bytes(await file.read())
        tam, sha = _save_pdf_to_store(pdf_path)
        row_id = _insert_processing(original_name, bank, empresa, RESULT_APP2DB["excel+cep"], tam, sha)

        # === XLSX temporal para empresa ===
        try:
            extractor_tmp = _get_extractor(bank)
            tmp_xlsx = Path(workdir) / (Path(original_name).stem + ".xlsx")
            await run_in_threadpool(extractor_tmp, str(pdf_path), str(tmp_xlsx))
            emp = _extract_empresa_from_xlsx(tmp_xlsx)
            if emp:
                _update_empresa(row_id, emp)
            try:
                if tmp_xlsx.exists():
                    tmp_xlsx.unlink()
            except Exception:
                pass
        except Exception:
            logger.exception("No se pudo generar XLSX temporal para leer 'empresa'")

        zip_str = await run_in_threadpool(_call_ceps, make_zip_with_ceps_for_bank, bank, str(pdf_path), str(workdir))
        zip_path = Path(zip_str)

        tam_out, sha_out, stored = _save_output_to_store(zip_path, "zip")
        _update_finish(row_id, EST_OK, salida_nombre=zip_path.name, salida_tamano=tam_out, salida_sha256=sha_out)

        tmp_path = _copy_for_response(stored); _schedule_cleanup(background_tasks, workdir, tmp_path)
        return FileResponse(path=tmp_path, filename=zip_path.name, media_type="application/zip")
    except Exception as e:
        _update_finish(row_id, EST_FAIL, str(e))
        return _as_error(e)

@app.post("/extract/bbva")
async def extract_bbva_only(background_tasks: BackgroundTasks, file: UploadFile = File(...), empresa: Optional[str] = Form(None)):
    from bbva_extractor import extract_bbva_to_xlsx
    bank = "bbva"
    original_name = file.filename or "estado.pdf"
    pdf_path: Optional[Path] = None
    row_id = 0
    try:
        workdir = _mk_workdir()
        pdf_path = workdir / original_name
        pdf_path.write_bytes(await file.read())
        tam, sha = _save_pdf_to_store(pdf_path)
        row_id = _insert_processing(original_name, bank, empresa, RESULT_APP2DB["excel"], tam, sha)

        out_xlsx = workdir / (Path(original_name).stem + "_bbva.xlsx")
        await run_in_threadpool(extract_bbva_to_xlsx, str(pdf_path), str(out_xlsx))

        # === Empresa desde XLSX ===
        try:
            emp = _extract_empresa_from_xlsx(out_xlsx)
            if emp:
                _update_empresa(row_id, emp)
        except Exception:
            logger.exception("No se pudo extraer/actualizar 'empresa' desde XLSX")

        tam_out, sha_out, stored = _save_output_to_store(out_xlsx, "xlsx")
        _update_finish(row_id, EST_OK, salida_nombre=out_xlsx.name, salida_tamano=tam_out, salida_sha256=sha_out)

        tmp_path = _copy_for_response(stored); _schedule_cleanup(background_tasks, workdir, tmp_path)
        return FileResponse(path=tmp_path, filename=out_xlsx.name,
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        _update_finish(row_id, EST_FAIL, str(e))
        return _as_error(e)

@app.post("/extract/bbva/with-cep")
async def extract_bbva_with_cep(background_tasks: BackgroundTasks, file: UploadFile = File(...), empresa: Optional[str] = Form(None)):
    bank = "bbva"
    original_name = file.filename or "estado.pdf"
    pdf_path: Optional[Path] = None
    row_id = 0
    try:
        make_zip_with_ceps_for_bank = _get_cep_service()
        if not make_zip_with_ceps_for_bank:
            return _as_error(Exception("Falta cep_service.py o Playwright."), status=501)

        workdir = _mk_workdir()
        pdf_path = workdir / original_name
        pdf_path.write_bytes(await file.read())
        tam, sha = _save_pdf_to_store(pdf_path)
        row_id = _insert_processing(original_name, bank, empresa, RESULT_APP2DB["excel+cep"], tam, sha)

        # === XLSX temporal para empresa ===
        try:
            extractor_tmp = _get_extractor(bank)
            tmp_xlsx = Path(workdir) / (Path(original_name).stem + ".xlsx")
            await run_in_threadpool(extractor_tmp, str(pdf_path), str(tmp_xlsx))
            emp = _extract_empresa_from_xlsx(tmp_xlsx)
            if emp:
                _update_empresa(row_id, emp)
            try:
                if tmp_xlsx.exists():
                    tmp_xlsx.unlink()
            except Exception:
                pass
        except Exception:
            logger.exception("No se pudo generar XLSX temporal para leer 'empresa'")

        zip_str = await run_in_threadpool(_call_ceps, make_zip_with_ceps_for_bank, bank, str(pdf_path), str(workdir))
        zip_path = Path(zip_str)

        tam_out, sha_out, stored = _save_output_to_store(zip_path, "zip")
        _update_finish(row_id, EST_OK, salida_nombre=zip_path.name, salida_tamano=tam_out, salida_sha256=sha_out)

        tmp_path = _copy_for_response(stored); _schedule_cleanup(background_tasks, workdir, tmp_path)
        return FileResponse(path=tmp_path, filename=zip_path.name, media_type="application/zip")
    except Exception as e:
        _update_finish(row_id, EST_FAIL, str(e))
        return _as_error(e)

# ---------------- Inbursa
@app.post("/extract/inbursa")
async def extract_inbursa_only(background_tasks: BackgroundTasks, file: UploadFile = File(...), empresa: Optional[str] = Form(None)):
    from inbursa_extractor import extract_inbursa_to_xlsx
    bank = "inbursa"
    original_name = file.filename or "estado.pdf"
    pdf_path: Optional[Path] = None
    row_id = 0
    try:
        workdir = _mk_workdir()
        pdf_path = workdir / original_name
        pdf_path.write_bytes(await file.read())
        tam, sha = _save_pdf_to_store(pdf_path)
        row_id = _insert_processing(original_name, bank, empresa, RESULT_APP2DB["excel"], tam, sha)

        out_xlsx = workdir / (pdf_path.stem + "_inbursa.xlsx")
        await run_in_threadpool(extract_inbursa_to_xlsx, str(pdf_path), str(out_xlsx))

        # === Empresa desde XLSX (si existiera en Inbursa) ===
        try:
            emp = _extract_empresa_from_xlsx(out_xlsx)
            if emp:
                _update_empresa(row_id, emp)
        except Exception:
            logger.exception("No se pudo extraer/actualizar 'empresa' desde XLSX")

        tam_out, sha_out, stored = _save_output_to_store(out_xlsx, "xlsx")
        _update_finish(row_id, EST_OK, salida_nombre=out_xlsx.name, salida_tamano=tam_out, salida_sha256=sha_out)

        tmp_path = _copy_for_response(stored); _schedule_cleanup(background_tasks, workdir, tmp_path)
        return FileResponse(path=tmp_path, filename=out_xlsx.name,
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        _update_finish(row_id, EST_FAIL, str(e))
        return _as_error(e)

@app.post("/extract/inbursa/with-cep")
async def extract_inbursa_with_cep(background_tasks: BackgroundTasks, file: UploadFile = File(...), empresa: Optional[str] = Form(None)):
    bank = "inbursa"
    original_name = file.filename or "estado.pdf"
    pdf_path: Optional[Path] = None
    row_id = 0
    try:
        make_zip_with_ceps_for_bank = _get_cep_service()
        if not make_zip_with_ceps_for_bank:
            return _as_error(Exception("Falta cep_service.py o Playwright."), status=501)

        workdir = _mk_workdir()
        pdf_path = workdir / original_name
        pdf_path.write_bytes(await file.read())
        tam, sha = _save_pdf_to_store(pdf_path)
        row_id = _insert_processing(original_name, bank, empresa, RESULT_APP2DB["excel+cep"], tam, sha)

        # === XLSX temporal para empresa ===
        try:
            extractor_tmp = _get_extractor(bank)
            tmp_xlsx = Path(workdir) / (Path(original_name).stem + ".xlsx")
            await run_in_threadpool(extractor_tmp, str(pdf_path), str(tmp_xlsx))
            emp = _extract_empresa_from_xlsx(tmp_xlsx)
            if emp:
                _update_empresa(row_id, emp)
            try:
                if tmp_xlsx.exists():
                    tmp_xlsx.unlink()
            except Exception:
                pass
        except Exception:
            logger.exception("No se pudo generar XLSX temporal para leer 'empresa'")

        zip_str = await run_in_threadpool(_call_ceps, make_zip_with_ceps_for_bank, bank, str(pdf_path), str(workdir))
        zip_path = Path(zip_str)

        tam_out, sha_out, stored = _save_output_to_store(zip_path, "zip")
        _update_finish(row_id, EST_OK, salida_nombre=zip_path.name, salida_tamano=tam_out, salida_sha256=sha_out)

        tmp_path = _copy_for_response(stored); _schedule_cleanup(background_tasks, workdir, tmp_path)
        return FileResponse(path=tmp_path, filename=zip_path.name, media_type="application/zip")
    except Exception as e:
        _update_finish(row_id, EST_FAIL, str(e))
        return _as_error(e)

# ==========================
#   HISTORIAL
# ==========================
def _parse_date_yyyy_mm_dd(date_str: str | None) -> Optional[datetime]:
    if not date_str: return None
    try: return datetime.strptime(date_str, "%Y-%m-%d")
    except Exception: return None

@app.get("/solicitudes")
def list_solicitudes(
    page: int = 1, page_size: int = 20,
    banco: str | None = None, empresa: str | None = None,
    resultado: str | None = None, estado: str | None = None,
    fecha_desde: str | None = None, fecha_hasta: str | None = None,
    q: str | None = None,
    user: dict = Depends(verify_jwt)
):
    if engine is None: raise HTTPException(status_code=503, detail="BD no configurada")
    page = max(1, page); page_size = max(1, min(200, page_size)); offset = (page - 1) * page_size

    conds = []
    id_usuario = user.get("id_usuario")
    conds.append(solicitudes.c.id_usuario == id_usuario)
    if banco:     conds.append(solicitudes.c.banco == banco.strip().lower())
    if resultado in {"xlsx","zip"}: conds.append(solicitudes.c.resultado == resultado)
    if estado in {EST_OK, EST_FAIL, EST_PROC}: conds.append(solicitudes.c.estado == estado)

    d1 = _parse_date_yyyy_mm_dd(fecha_desde); d2 = _parse_date_yyyy_mm_dd(fecha_hasta)
    if d1: conds.append(solicitudes.c.solicitado_en >= d1)
    if d2: conds.append(solicitudes.c.solicitado_en < d2.replace(hour=23, minute=59, second=59, microsecond=999999))

    if empresa: conds.append(solicitudes.c.empresa.ilike(f"%{empresa.strip()}%"))
    if q:
        like = f"%{q.strip()}%"
        conds.append((solicitudes.c.archivo_nombre.ilike(like)) | (solicitudes.c.empresa.ilike(like)))

    where_all = and_(*conds) if conds else None

    with engine.begin() as conn:
        total = conn.execute(
            select(func.count()).select_from(solicitudes).where(where_all)
        ).scalar_one()
        stmt = (
            select(
                solicitudes.c.id, solicitudes.c.solicitado_en, solicitudes.c.archivo_nombre,
                solicitudes.c.archivo_tamano, solicitudes.c.archivo_sha256, solicitudes.c.banco,
                solicitudes.c.empresa, solicitudes.c.resultado, solicitudes.c.estado, solicitudes.c.error,
                solicitudes.c.salida_nombre, solicitudes.c.salida_tamano, solicitudes.c.salida_sha256,
                solicitudes.c.id_usuario, solicitudes.c.nombre_usuario
            ).where(where_all)
        )
        rows = conn.execute(stmt.order_by(desc(solicitudes.c.id)).offset(offset).limit(page_size)).mappings().all()

    return {"total": total, "page": page, "page_size": page_size, "items": [dict(r) for r in rows]}

@app.get("/solicitudes/export")
def export_solicitudes_csv(
    banco: str | None = None, empresa: str | None = None,
    resultado: str | None = None, estado: str | None = None,
    fecha_desde: str | None = None, fecha_hasta: str | None = None,
    q: str | None = None,
):
    if engine is None: raise HTTPException(status_code=503, detail="BD no configurada")

    conds = []
    if banco: conds.append(solicitudes.c.banco == banco.strip().lower())
    if resultado in {"xlsx","zip"}: conds.append(solicitudes.c.resultado == resultado)
    if estado in {EST_OK, EST_FAIL, EST_PROC}: conds.append(solicitudes.c.estado == estado)

    d1 = _parse_date_yyyy_mm_dd(fecha_desde); d2 = _parse_date_yyyy_mm_dd(fecha_hasta)
    if d1: conds.append(solicitudes.c.solicitado_en >= d1)
    if d2: conds.append(solicitudes.c.solicitado_en < d2.replace(hour=23, minute=59, second=59, microsecond=999999))

    if empresa: conds.append(solicitudes.c.empresa.ilike(f"%{empresa.strip()}%"))
    if q:
        like = f"%{q.strip()}%"
        conds.append((solicitudes.c.archivo_nombre.ilike(like)) | (solicitudes.c.empresa.ilike(like)))

    where_all = and_(*conds) if conds else None

    with engine.begin() as conn:
        stmt = (
            select(
                solicitudes.c.id, solicitudes.c.solicitado_en, solicitudes.c.archivo_nombre,
                solicitudes.c.archivo_tamano, solicitudes.c.archivo_sha256, solicitudes.c.banco,
                solicitudes.c.empresa, solicitudes.c.resultado, solicitudes.c.estado, solicitudes.c.error,
            ).where(where_all) if where_all else
            select(
                solicitudes.c.id, solicitudes.c.solicitado_en, solicitudes.c.archivo_nombre,
                solicitudes.c.archivo_tamano, solicitudes.c.archivo_sha256, solicitudes.c.banco,
                solicitudes.c.empresa, solicitudes.c.resultado, solicitudes.c.estado, solicitudes.c.error,
            )
        )
        rows = conn.execute(stmt.order_by(desc(solicitudes.c.id))).mappings().all()

    buf = io.StringIO(newline="")
    w = csv.writer(buf)
    w.writerow(["id","fecha","archivo_nombre","tamano","sha256","banco","empresa","resultado","estado","error"])
    for r in rows:
        w.writerow([
            r["id"],
            r["solicitado_en"].strftime("%Y-%m-%d %H:%M:%S") if r["solicitado_en"] else "",
            r["archivo_nombre"] or "", r["archivo_tamano"] or "", r["archivo_sha256"] or "",
            r["banco"] or "", r["empresa"] or "", r["resultado"] or "", r["estado"] or "",
            (r["error"] or "").replace("\n"," ").replace("\r"," "),
        ])
    buf.seek(0)

    filename = f"solicitudes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(iter([buf.read()]), media_type="text/csv", headers=headers)

# =========================================================
#   Ver PDF fuente (inline en el navegador)
# =========================================================
@app.get("/file/{solicitud_id}")
def view_original_pdf(solicitud_id: int):
    if engine is None:
        raise HTTPException(status_code=503, detail="BD no configurada")

    # Buscar SHA y nombre del archivo en BD
    with engine.begin() as conn:
        row = conn.execute(
            select(
                solicitudes.c.id,
                solicitudes.c.archivo_nombre,
                solicitudes.c.archivo_sha256,
            ).where(solicitudes.c.id == solicitud_id)
        ).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")

    sha = (row["archivo_sha256"] or "").strip()
    if not sha:
        # Para registros viejos que no guardaron SHA
        raise HTTPException(status_code=404, detail="Solicitud sin PDF asociado")

    src = INPUT_DIR / f"{sha}.pdf"
    if not src.exists():
        raise HTTPException(status_code=404, detail="PDF no disponible en el repositorio")

    # Mostrar inline en el navegador
    filename = row["archivo_nombre"] or f"solicitud_{solicitud_id}.pdf"
    headers = {"Content-Disposition": f'inline; filename="{filename}"'}

    return FileResponse(
        path=str(src),
        headers=headers,
        media_type="application/pdf",
        filename=filename,  # por si el navegador decide descargar
    )

# =========================================================
#   Descarga desde historial (SIN reprocesar)
# =========================================================
@app.get("/download/{solicitud_id}")
async def download_solicitud(solicitud_id: int, background_tasks: BackgroundTasks):
    if engine is None:
        raise HTTPException(status_code=503, detail="BD no configurada")

    with engine.begin() as conn:
        row = conn.execute(
            select(
                solicitudes.c.id,
                solicitudes.c.archivo_nombre,
                solicitudes.c.resultado,
                solicitudes.c.estado,
                solicitudes.c.salida_nombre,
                solicitudes.c.salida_sha256,
            ).where(solicitudes.c.id == solicitud_id)
        ).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="No encontrada")
    if row["estado"] != "ok":
        raise HTTPException(status_code=409, detail="No disponible: la extracción no fue completada")

    sha_out = row["salida_sha256"]
    if not sha_out:
        raise HTTPException(status_code=404, detail="No hay salida asociada a la solicitud")

    # Resuelve extensión desde lo guardado en BD
    ext = "xlsx" if (row["resultado"] or "").lower() == "xlsx" else "zip"
    src = OUTPUT_DIR / f"{sha_out}.{ext}"
    if not src.exists():
        raise HTTPException(status_code=404, detail="Archivo no disponible en el repositorio")

    # Nombre de descarga: usa el nombre guardado; si no existe, deriva del nombre del PDF
    default_pdf = "solicitud_" + str(row["id"]) + ".pdf"
    stem = Path(row["archivo_nombre"] or default_pdf).stem
    filename = row["salida_nombre"] or (stem + (".xlsx" if ext == "xlsx" else "_ceps.zip"))

    tmp_path = _copy_for_response(src)
    _schedule_cleanup(background_tasks, tmp_path)

    return FileResponse(
        path=tmp_path,
        filename=filename,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if ext == "xlsx" else "application/zip"
        ),
    )
